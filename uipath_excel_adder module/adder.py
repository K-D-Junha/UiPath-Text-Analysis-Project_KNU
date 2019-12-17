import openpyxl
import nltk
import time
from openpyxl import load_workbook
from nltk.sentiment.vader import SentimentIntensityAnalyzer #감정분석
import re #정규식 사용
from datetime import datetime, timedelta #시간계산 datetime(날짜), time,delta(시간의 차)
from pytz import timezone #시간대 변경
from googletrans import Translator#구글 번역
from json import JSONDecodeError


def open_workbook( file = None): 
  workbook = load_workbook(file, data_only=True)#엑셀 열기
  return workbook

def clean_str(text):
  try:
    pattern = '([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)' # E-mail제거
    text = re.sub(pattern=pattern, repl='', string=text)
    pattern = '(http|ftp|https)://(?:[-\w.]|(?:%[\da-fA-F]{2}))+' # URL제거
    text = re.sub(pattern=pattern, repl='', string=text)
    pattern = '<[^>]*>'         # HTML 태그 제거
    text = re.sub(pattern=pattern, repl='', string=text)
    #pattern = '[^\w\s]'         # 특수기호제거
    #text = re.sub(pattern=pattern, repl='', string=text)
    pattern = re.compile("["
                          u"\U0001F600-\U0001F64F"  # emoticons
                          u"\U0001F300-\U0001F5FF"  # symbols & pictographs
                          u"\U0001F680-\U0001F6FF"  # transport & map symbols
                          u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
                          "]+", flags=re.UNICODE)
    text = re.sub(pattern=pattern, repl='', string=text)
  except TypeError:
    test = 'neutral'
  return text     

def get_comments(sheet,startrow=1) :
  result = []
  row_num = sheet.max_row#시트 줄수
  col_num = sheet.max_column#시트 열수

  if row_num<startrow:
    return result

  for r in range(startrow, row_num+1) :
    clean_comment = clean_str(sheet.cell(r,3).value)
    if clean_comment is None: #empty set
      clean_comment = 'neutral'
    result.append(clean_comment)

  return result

def translate_com(comments = None):
  result = []
  for i in range(0,len(comments)) :
    try:
      translator = Translator(service_urls=[
      'translate.google.com',
      'translate.google.co.kr',
      'translate.google.co.jp',
      'translate.google.co.uk',
      ])

      #lang = detect(comments[i])
      #print(i,translator.detect(comments[i]).lang, comments[i])
      if(translator.detect(comments[i]).lang == 'en'):
        result.append(comments[i])
      else:
        time.sleep(1)
        trs_comment = translator.translate(comments[i])
        print(comments[i], "\ntranslated : ",trs_comment.text)
        result.append(trs_comment.text)
    except (AttributeError, TypeError):
      result.append('neutral')
    except JSONDecodeError:
      print("JSONERROR :", comments[i])
      result.append(comments[i])
  
  return result

def analyze_emotion(trans_list = None) :#점수 매기기
  nltk.download("book")
  nltk.download('vader_lexicon')
  sid = SentimentIntensityAnalyzer()

  result = []
  for i in trans_list:
    score = sid.polarity_scores(i)
    result.append(score)
    
  return result

def get_like(sheet = None, startrow = 1) : #int + 답변의 내용에서 좋아요 숫자를 구하기(정규표현식)
  desc = re.compile('\d*\d')
  result = []

  for i in range(startrow,sheet.max_row+1) :
    find_like = desc.findall(sheet.cell(i,4).value)

    if len(find_like) is None:
      result.append('0')
    else :
      #print(find_like)
      if len(find_like) == 1 :
        result.append('0')
      elif len(find_like) == 2:
        #print(int(find_like[1])-1)
        result.append(str(int(find_like[1])-1))    
      else : result.append('0')        
  return result


def get_realtime(sheet = None, startrow = 1):#시간계산
  #년,달,주,일,시간,분

  minute_desc = re.compile('\d*\d분')
  hour_desc = re.compile('\d*\d시간')
  day_desc = re.compile('\d*\d일')
  week_desc = re.compile('\d*\d주')
  month_desc = re.compile('\d*\d개월')
  year_desc = re.compile('\d*\d년')
  desc = re.compile('\d*\d')

  #시간 출력 포맷
  format = "%Y-%m-%d %H:%M:%S"

  #현재시간
  UTC = datetime.now(timezone('UTC'))
  KST = datetime.now(timezone('Asia/Seoul'))

  #print(KST)

  result = []
  # ~전 토큰으로 자름 return객체 = list
  for i in range(startrow,sheet.max_row+1) :
    minute_ago = minute_desc.findall(sheet.cell(i,2).value)
    hour_ago = hour_desc.findall(sheet.cell(i,2).value)
    day_ago = day_desc.findall(sheet.cell(i,2).value)
    week_ago = week_desc.findall(sheet.cell(i,2).value)
    month_ago = month_desc.findall(sheet.cell(i,2).value)
    year_ago = year_desc.findall(sheet.cell(i,2).value)
    #print(minute_ago,hour_ago,day_ago,week_ago,month_ago,year_ago)
    
    # 길이 != 0 -> 현재시간 - 차이 시간 = return객체 = datetime -> str으로 변경
    if len(minute_ago) != 0:
      min_out = int((desc.findall(minute_ago[0]))[0])
      rtime = KST-timedelta(minutes = min_out)
      str_rtime = rtime.strftime(format)
      #print(str_rtime)
      result.append(str_rtime)
      
    elif len(hour_ago) != 0:
      hour_out = int((desc.findall(hour_ago[0]))[0])
      rtime = KST-timedelta(hours = hour_out)
      str_rtime = rtime.strftime(format)
      #print(str_rtime)
      result.append(str_rtime)

    elif len(day_ago) != 0:
      day_out = int((desc.findall(day_ago[0]))[0])
      rtime = KST-timedelta(days = day_out)
      str_rtime = rtime.strftime(format)
      #print(str_rtime)
      result.append(str_rtime)

    elif len(week_ago) != 0:
      week_out = int((desc.findall(week_ago[0]))[0])
      rtime = KST-timedelta(days = week_out*7)
      str_rtime = rtime.strftime(format)
      #print(str_rtime)
      result.append(str_rtime)

    elif len(month_ago) != 0:
      month_out = int((desc.findall(month_ago[0]))[0])
      rtime = KST-timedelta(days = month_out*30)
      str_rtime = rtime.strftime(format)
      #print(str_rtime)
      result.append(str_rtime)

    elif len(year_ago) != 0:
      year_out = int((desc.findall(year_ago[0]))[0])
      rtime = KST-timedelta(days = year_out*365)
      str_rtime = rtime.strftime(format)
      #print(str_rtime)
      result.append(str_rtime)
    
    else :
      result.append(KST.strftime(format))
  
  return result

###############################################################################

def add_score2excel(file = None,sheet = None,startrow=1): #파일명, 시작row
  endtype = {1 : "success", 2:'fail', 3:'no change'}  
  if file is None and sheet is None:
    return endtype[2]
  if startrow < 1 :
    startrow = 1
  
  book = open_workbook(file)
  sheet1 = book[sheet] 
  
  time_list = get_realtime(sheet1,startrow) # 절대 시간 리스트
  like_list = get_like(sheet1,startrow) # like 수 리스트
  comment_list = get_comments(sheet1,startrow)# 원본 코멘트 리스트
  trans_list = translate_com(comment_list)# 번역된 리스트
  score_list = analyze_emotion(trans_list) # 점수 리스트

  #print(len(comment_list),len(trans_list),len(score_list),len(like_list),len(time_list))

  if(len(score_list)==0):
      return endtype[3]
  sheet1_row = sheet1.max_row
  #print(len(score_list))

  if startrow == 1 :
    sheet1.insert_rows(1)
    sheet1.cell(1,1).value = 'id'
    sheet1.cell(1,2).value = 'time'
    sheet1.cell(1,3).value = 'original'
    sheet1.cell(1,4).value = 'translated'
    sheet1.cell(1,5).value = 'like'
    sheet1.cell(1,6).value = 'neg'
    sheet1.cell(1,7).value = 'neu'
    sheet1.cell(1,8).value = 'pos'
    sheet1.cell(1,9).value = 'comp'
    writerow = startrow+1
  
  else :
    writerow = startrow

  for i in range(0,len(score_list)):
    #print(writerow ,score_list[i])
    sheet1.cell(writerow,2).value = time_list[i]
    sheet1.cell(writerow,4).value = trans_list[i]
    sheet1.cell(writerow,5).value = like_list[i]
    sheet1.cell(writerow,6).value = score_list[i]['neg']
    sheet1.cell(writerow,7).value = score_list[i]['neu']
    sheet1.cell(writerow,8).value = score_list[i]['pos']
    sheet1.cell(writerow,9).value = score_list[i]['compound']
    writerow+=1
    
  book.save(file)
  return endtype[1]