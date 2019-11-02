from openpyxl import load_workbook
from nltk.sentiment.vader import SentimentIntensityAnalyzer
import nltk

nltk.download('vader_lexicon')
sid = SentimentIntensityAnalyzer()
'''
def analyze(file = None,temp =0) :
    if file == None :
        return 'There is No File!!!\n'
    
    load_wb = load_workbook(file, data_only=True)

    load_ws = load_wb['Sheet1']
    result = []
   
    return load_ws['C3'].value +"\n" + str(sid.polarity_scores(load_ws['C3'].value))
	
def test(one =0):
	return one+100

def test2(one=0,two=0):
	haha = test(one)
	return haha
	
'''

def open_workbook(file = None): 
  workbook = load_workbook(file, data_only=True)
  return workbook

def get_all_comments(workbook = None,start=2) :
  worksheet = workbook['Sheet1']
  result = []

  row_num = worksheet.max_row
  col_num = worksheet.max_column
  if row_num<start:
    return result

  for r in range(start, row_num+1) :
    #result.append(worksheet['C'+str(r)].value)
    #print(start)
    #print(worksheet['C'+str(r)].value)
    result.append(worksheet.cell(r,3).value)

  return result

def analyze_emotion(comment_list = None) :
  result = []
  for i in comment_list:
    score = sid.polarity_scores(i)
    result.append(score)
    
  return result


def add_score2excel(file = None,startrow=2): #파일명, 시작row, sheet2에 점수저장
  endtype = {1 : "success", 2:'fail', 3:'nothing to analyze'}
  if file == None :
    return endtype[2]
  if startrow < 2 :
    startrow = 2
  
  book = open_workbook(file)
  sheet1 = book['Sheet1']
  
  comment_list = get_all_comments(book,startrow)
  score_list = analyze_emotion(comment_list)

  if(len(score_list)==0):
      return endtype[3]
    

  if startrow==2: #한번도 감정분석을 하지 않아서 목차에 아무것도 없을 때
    #sheet1.cell(1,4).value = 'comment_row'
    sheet1.cell(1,4).value = 'neg'
    sheet1.cell(1,5).value = 'neu'
    sheet1.cell(1,6).value = 'pos'
    sheet1.cell(1,7).value = 'compound'

  sheet1_row = sheet1.max_row
  #print(len(score_list))

  writerow = startrow
  for i in range(0,len(score_list)):
    #print(writerow ,score_list[i])
    sheet1.cell(writerow,4).value = score_list[i]['neg']
    sheet1.cell(writerow,5).value = score_list[i]['neu']
    sheet1.cell(writerow,6).value = score_list[i]['pos']
    sheet1.cell(writerow,7).value = score_list[i]['compound']
    writerow+=1

  book.save(file)
  return endtype[1]
  


#testline
'''
file = "C:\datasource\Samsung.xlsx"
print(add_score2excel(file,2))
'''
